import pflib.pf as pf
import pandas as pd
import numpy as np
import os
import importlib
from openpyxl import load_workbook
import datetime
from experiment_config import experiment_path, chosen_experiment
import math

spec = importlib.util.spec_from_file_location(chosen_experiment, experiment_path)
config = importlib.util.module_from_spec(spec)
spec.loader.exec_module(config)
learning_config = config.learning_config


def define_PV_controls(app):
    '''
    define cos(phi)/Q(P) control for PVs
    '''

    # Get folder where to create new QP char object
    o_QPCurves_IntPrjfolder = app.GetProjectFolder('qpc')

    # Clear QP char folder
    for o in o_QPCurves_IntPrjfolder.GetContents():
        o.Delete()

    # Clear Capability Curve Folder (for Q and P limits)
    o_QlimCurve_IntPrjfolder = app.GetProjectFolder('mvar')
    for o in o_QlimCurve_IntPrjfolder.GetContents():
        o.Delete()

    # Create cosphi(P) char object and set attributes
    o_IntcosphiPcurve = o_QPCurves_IntPrjfolder.CreateObject('IntQpcurve', 'QP acting as cosphi(P) char')
    o_IntcosphiPcurve.SetAttribute('inputmod', 1)
    o_IntcosphiPcurve.SetAttribute('Ppu',
                                   [0, 0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12, 0.13,
                                    0.14, 0.15,
                                    0.16, 0.17, 0.18, 0.19, 0.2, 0.21, 0.22, 0.23, 0.24, 0.25, 0.26, 0.27, 0.28, 0.29,
                                    0.3, 0.31,
                                    0.32, 0.33, 0.34, 0.35, 0.36, 0.37, 0.38, 0.39, 0.4, 0.41, 0.42, 0.43, 0.44, 0.45,
                                    0.46,
                                    0.47, 0.48, 0.49, 0.5, 0.51, 0.52, 0.53, 0.54, 0.55, 0.56, 0.57, 0.58, 0.59, 0.6,
                                    0.61, 0.62,
                                    0.63, 0.64, 0.65, 0.66, 0.67, 0.68, 0.69, 0.7, 0.71, 0.72, 0.73, 0.74, 0.75, 0.76,
                                    0.77,
                                    0.78, 0.79, 0.8, 0.81, 0.82, 0.83, 0.84, 0.85, 0.86, 0.87, 0.88, 0.89, 0.9, 0.91,
                                    0.92, 0.93,
                                    0.94, 0.95, 0.96, 0.97, 0.98, 0.99, 1
                                    ])
    o_IntcosphiPcurve.SetAttribute('Qpu',
                                   [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                                    0, 0, 0,
                                    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0.0633406, 0.08971208,
                                    0.11004031, 0.12725592, 0.14249228, 0.15632983, 0.1691129, 0.18106551, 0.19234309,
                                    0.20305866, 0.21329743, 0.22312553, 0.23259549, 0.24174985, 0.25062362, 0.25924607,
                                    0.26764189, 0.2758322, 0.28383518, 0.29166667, 0.2993405, 0.3068689, 0.31426269,
                                    0.32153154,
                                    0.32868411, 0.33572819, 0.34267085, 0.34951849, 0.35627693, 0.36295153, 0.36954718,
                                    0.37606838, 0.38251928, 0.38890373, 0.39522529, 0.40148728, 0.40769277, 0.41384466,
                                    0.41994564, 0.42599822, 0.43200477, 0.43796754, 0.44388861, 0.44976996, 0.45561346,
                                    0.46142089, 0.46719391, 0.47293413, 0.47864305, 0.4843221])

    # Create Q(P) char object and set attributes
    o_IntQpcurve = o_QPCurves_IntPrjfolder.CreateObject('IntQpcurve', 'General PQ char')
    o_IntQpcurve.SetAttribute('inputmod', 1)
    o_IntQpcurve.SetAttribute('Ppu', [0, 0.5, 1])
    o_IntQpcurve.SetAttribute('Qpu', [0, 0, -0.338])

    # Create disfunctional Q(P) char object and set attributes
    o_brokenIntQpcurve = o_QPCurves_IntPrjfolder.CreateObject('IntQpcurve', 'Broken PQ char')
    o_brokenIntQpcurve.SetAttribute('inputmod', 1)
    o_brokenIntQpcurve.SetAttribute('Ppu', [0, 0.5, 1])
    o_brokenIntQpcurve.SetAttribute('Qpu', [0, 0, 0])

    # Create wrong Q(P) char object and set attributes (inverse curve)
    o_wrongIntcosphiPcurve = o_QPCurves_IntPrjfolder.CreateObject('IntQpcurve', 'Wrong QP acting as cosphi(P) char')
    o_wrongIntcosphiPcurve.SetAttribute('inputmod', 1)
    o_wrongIntcosphiPcurve.SetAttribute('Ppu',
                                        [0, 0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12, 0.13,
                                         0.14, 0.15,
                                         0.16, 0.17, 0.18, 0.19, 0.2, 0.21, 0.22, 0.23, 0.24, 0.25, 0.26, 0.27, 0.28,
                                         0.29,
                                         0.3, 0.31,
                                         0.32, 0.33, 0.34, 0.35, 0.36, 0.37, 0.38, 0.39, 0.4, 0.41, 0.42, 0.43, 0.44,
                                         0.45,
                                         0.46,
                                         0.47, 0.48, 0.49, 0.5, 0.51, 0.52, 0.53, 0.54, 0.55, 0.56, 0.57, 0.58, 0.59,
                                         0.6,
                                         0.61, 0.62,
                                         0.63, 0.64, 0.65, 0.66, 0.67, 0.68, 0.69, 0.7, 0.71, 0.72, 0.73, 0.74, 0.75,
                                         0.76,
                                         0.77,
                                         0.78, 0.79, 0.8, 0.81, 0.82, 0.83, 0.84, 0.85, 0.86, 0.87, 0.88, 0.89, 0.9,
                                         0.91,
                                         0.92, 0.93,
                                         0.94, 0.95, 0.96, 0.97, 0.98, 0.99, 1
                                         ])
    o_wrongIntcosphiPcurve.SetAttribute('Qpu',
                                        list(-np.array(
                                            [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                                             0, 0, 0,
                                             0, 0, 0,
                                             0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0.0633406,
                                             0.08971208,
                                             0.11004031, 0.12725592, 0.14249228, 0.15632983, 0.1691129, 0.18106551,
                                             0.19234309,
                                             0.20305866, 0.21329743, 0.22312553, 0.23259549, 0.24174985, 0.25062362,
                                             0.25924607,
                                             0.26764189, 0.2758322, 0.28383518, 0.29166667, 0.2993405, 0.3068689,
                                             0.31426269,
                                             0.32153154,
                                             0.32868411, 0.33572819, 0.34267085, 0.34951849, 0.35627693, 0.36295153,
                                             0.36954718,
                                             0.37606838, 0.38251928, 0.38890373, 0.39522529, 0.40148728, 0.40769277,
                                             0.41384466,
                                             0.41994564, 0.42599822, 0.43200477, 0.43796754, 0.44388861, 0.44976996,
                                             0.45561346,
                                             0.46142089, 0.46719391, 0.47293413, 0.47864305, 0.4843221])))

    return o_IntcosphiPcurve, o_IntQpcurve, o_brokenIntQpcurve, o_wrongIntcosphiPcurve


def place_PVs(app, o_ElmNet, o_ChaTime, loads_by_type, PV_apparent_power=0.005, type_QDSL_model=None):
    '''
    Place photvoltaics next to every load, assign control/capability curve &  charactersitic and scale their output
    to the consumption of the load they are attached to so as it yields about the yearly consumption of the load
    '''

    type_qdsl_model = import_QDSL_type('cosphi(P)')
    curves = define_PV_controls(app)

    o_QlimCurve_IntPrjfolder = app.GetProjectFolder('mvar')
    o_IntQlim = o_QlimCurve_IntPrjfolder.SearchObject('Capability Curve')
    if o_IntQlim is None or o_IntQlim.loc_name != 'Capability Curve':
        o_IntQlim = o_QlimCurve_IntPrjfolder.CreateObject('IntQlim', 'Capability Curve')
        o_IntQlim.SetAttribute('cap_Ppu', [0, 1])
        o_IntQlim.SetAttribute('cap_Qmnpu',
                               [0, -0.436])  # operational limits following the standard; max cosphi = 0.9
        o_IntQlim.SetAttribute('cap_Qmxpu', [0, 0.436])
        o_IntQlim.SetAttribute('inputmod', 1)

    for load in loads_by_type['regular_loads']:
        o_ElmLod = load[0]
        load_cubicle = o_ElmLod.bus1  # elements are connected to terminals via cubicles in powerfactory

        o_ElmTerm = load_cubicle.cterm

        o_StaCubic = o_ElmTerm.CreateObject('StaCubic',
                                            'Cubicle_' + o_ElmLod.loc_name.split(' ')[0] + ' SGen ' +
                                            o_ElmLod.loc_name.split(' ')[2])
        o_Elm = o_ElmNet.CreateObject('ElmGenstat',
                                      o_ElmLod.loc_name.split(' ')[0] + ' SGen ' + o_ElmLod.loc_name.split(' ')[
                                          2] + ' @ ' + o_ElmLod.bus1.cterm.loc_name
                                      )
        o_Elm.SetAttribute('bus1', o_StaCubic)
        o_Elm.SetAttribute('sgn', PV_apparent_power)
        o_Elm.cCategory = 'Photovoltaic'
        o_Elm.outserv = 1  # deactivate all PVs at first and then activate random ones during simulation

        if not config.QDSL_models_available or config.number_of_broken_devices_and_type[1] == 'PV':
            o_Elm.pgini = o_Elm.sgn * 0.9 * (o_ElmLod.plini / 0.004)  # scale with yearly consumption of load
            pf.set_referenced_characteristics(o_Elm, 'pgini', o_ChaTime)  # set characteristic for inserted PV
            o_Elm.SetAttribute('av_mode', 'qpchar')  # Control activated
            o_Elm.SetAttribute('pQPcurve', curves[config.control_curve_choice])  # Control assigned

            o_Elm.SetAttribute('Pmax_ucPU', 1)  # set the operational limits for the PV
            o_Elm.SetAttribute('pQlimType', o_IntQlim)
        elif config.QDSL_models_available:
            o_Elm.pgini = o_Elm.sgn * 0.9 * (
                        o_ElmLod.plini / 0.004)  # scale with yearly consumption of load > is reset later bc of QDSL model usage
            o_Elm.SetAttribute('pQPcurve', curves[config.control_curve_choice])  # Control assigned
            # insert QDSL model to do convergence loop when control is active
            o_ElmQDSLmodel = o_ElmNet.CreateObject('ElmQdsl',
                                                   'QDSLmodel_' + o_ElmLod.loc_name.split(' ')[0] + ' PV ' +
                                                   o_ElmLod.loc_name.split(' ')[
                                                       2])
            o_ElmQDSLmodel.typ_id = type_qdsl_model
            o_ElmQDSLmodel.SetAttribute('e:initVals', [0.0, 0.0])
            o_ElmQDSLmodel.SetAttribute('e:objectsLdf', [o_Elm, o_Elm])
            o_Elm.i_scale = 0

    return curves


def place_home_or_work_EVCS(loads_by_type, loads, type, o_ElmNet, type_QDSL_model=None):
    EV_charging_stations = []
    if type not in ['Home', 'Work']:
        print('invalid type, valid types are: Home, Work')
        return EV_charging_stations

    for number, load in enumerate(loads):
        o_ElmLod = load[0]
        load_cubicle = o_ElmLod.bus1  # elements are connected to terminals via cubicles in powerfactory
        o_ElmTerm = load_cubicle.cterm
        o_StaCubic = o_ElmTerm.CreateObject('StaCubic',
                                            'Cubicle_' + o_ElmLod.loc_name.split(' ')[0] + ' EVCS ' +
                                            o_ElmLod.loc_name.split(' ')[2])
        o_ElmEVCS = o_ElmNet.CreateObject('ElmLod',
                                          o_ElmLod.loc_name.split(' ')[0] + ' EVCS ' + o_ElmLod.loc_name.split(' ')[
                                              2] + ' @ ' + o_ElmLod.bus1.cterm.loc_name)
        o_ElmEVCS.SetAttribute('bus1', o_StaCubic)

        parameters = loads_by_type['EV_charging_stations'][type][
            number % len(loads_by_type['EV_charging_stations'][type])]

        if not config.QDSL_models_available or config.number_of_broken_devices_and_type[1] == 'PV':
            o_ElmEVCS.SetAttribute('typ_id', parameters[2])
            pf.set_referenced_characteristics(o_ElmEVCS, 'plini',
                                              parameters[0][0])  # set characteristic for inserted EVCS
            o_ElmEVCS.plini = parameters[0][1]
            pf.set_referenced_characteristics(o_ElmEVCS, 'qlini',
                                              parameters[1][0])  # set characteristic for inserted EVCS
            o_ElmEVCS.qlini = parameters[1][1]
        elif config.QDSL_models_available:
            pf.set_referenced_characteristics(o_ElmEVCS, 'plini',
                                              parameters[0][0])  # set characteristic for inserted EVCS
            o_ElmEVCS.plini = parameters[0][1]
            # insert QDSL model to do convergence loop when control is active
            o_ElmQDSLmodel = o_ElmNet.CreateObject('ElmQdsl',
                                                   'QDSLmodel_' + o_ElmLod.loc_name.split(' ')[0] + ' EVCS ' +
                                                   o_ElmLod.loc_name.split(' ')[
                                                       2])
            o_ElmQDSLmodel.typ_id = type_QDSL_model
            o_ElmQDSLmodel.SetAttribute('e:initVals', [0.0, 0.0, 1.05, 0.95, 0])
            o_ElmQDSLmodel.SetAttribute('e:objectsLdf', [o_ElmEVCS, o_ElmTerm])
            o_ElmEVCS.i_scale = 0

        o_ElmEVCS.outserv = 1  # deactivate all EVCSs at first and then activate random ones during simulation

        EV_charging_stations.append(o_ElmEVCS)

    return EV_charging_stations


def import_QDSL_type(control_algorithm):
    target_folder = pf.app.GetProjectFolder('blk')

    if control_algorithm == 'p_of_u':
        path = os.path.join(config.grid_data_folder, 'QDSLModels.dz')
        folder = pf.app.ImportDz(target_folder, path)[1][0]
        type = folder.GetContents('p_of_u')[0]
    elif control_algorithm == 'cosphi(P)':
        path = os.path.join(config.grid_data_folder, 'QDSLModels.dz')
        folder = pf.app.ImportDz(target_folder, path)[1][0]
        type = folder.GetContents('cosphi(P)')[0]
    else:
        type = None

    return type


def place_EVCS(o_ElmNet, loads_by_type):
    '''
        Place EV charging stations next to every load and assign charactersitic (according to load type)
     '''

    homes = [i for i in loads_by_type['regular_loads'] if i[1].loc_name[0] == 'H']
    companies = [i for i in loads_by_type['regular_loads'] if i[1].loc_name[0] == 'G']

    type_qdsl_model = import_QDSL_type('p_of_u')
    EV_charging_stations = place_home_or_work_EVCS(loads_by_type, homes, 'Home', o_ElmNet, type_qdsl_model)
    EV_charging_stations = EV_charging_stations + place_home_or_work_EVCS(loads_by_type, companies, 'Work',
                                                                          o_ElmNet,
                                                                          type_qdsl_model)  # making sure a work charging station is palced next to a company and a home charging station next to a home

    return EV_charging_stations


def utf8_complaint_naming(o_ElmNet):
    '''
    Check for UTF-8 correct naming of elements (important for reading the result file into a dataframe)
    '''

    l_objects = o_ElmNet.GetContents(1)  # get all network elements
    not_utf = []
    for o in l_objects:
        string = o.loc_name
        if len(string.encode('utf-8')) == len(string):  # check if name is UTF-8
            # print ("string is UTF-8, length %d bytes" % len(string))
            continue
        else:
            print("string is not UTF-8")
            count = 0
            not_utf.append(o)
            for l in string:
                if len(l.encode('utf-8')) > 1:  # replace not UTF-8 character with _
                    new = string.replace(l, '_', count)
                    string = new
                count += 1
            o.loc_name = string
    print('%d element names changed to match UTF-8 format' % len(not_utf))

    return 0

def convert_to_unix_timestamp(pd_timestamp):

    unix_timestamp = (pd_timestamp - pd.Timestamp("1970-01-01", tz='utc')) // pd.Timedelta('1ns')

    return unix_timestamp

def create_characteristics(element, chars_dict, sim_setting=config.sim_setting, element_type='load', data=None, setup=None, extract_profiles=False, pv_input=None):
    if sim_setting in ['ERIGrid_phase_1', 'phase1']:
        folder = 'ERIGrid_Profiles_phase1'
    elif sim_setting in ['ERIGrid_phase_2', 'phase2']:
        folder = 'ERIGrid_Profiles_phase2'
        #profiles_path = os.path.join(config.grid_data_folder, folder)
        import raw_data_generation.input.ERIGrid_Profiles_phase2.load_script as load_script
    elif sim_setting in ['stmk']:
        folder = 'ENS_' + learning_config['setup_chosen']['stmk']
    else:
        print('Undefined simulation setting!')

    if type(data) is list:
        sensor_data = data[0]   #included in profiles, just here for optional use
        profiles = data[1]
    elif type(data) is pd.DataFrame:
        loads = [load.split('_')[0] for load in list(data.columns[0::2])]
        profiles = data
    elif sim_setting in ['stmk']:
        voltage_folder = folder + '\\UW_20kV_Abgang'
        voltage_data = pd.read_excel(os.path.join(config.grid_data_folder, voltage_folder, 'UW_20kV_Abgang_U.xlsx'))

        if learning_config['setup_chosen']['stmk'] == 'Gleinz': Un = 20.5
        elif learning_config['setup_chosen']['stmk'] == 'Neudau': Un = 20.6
        else: Un = 20
        voltage_data.insert(loc=2, column='mean voltage p.u.', value=(voltage_data.iloc[:, 2:]).mean(axis=1)*math.sqrt(3)/Un)

        pv_folder = folder + '\\NAP'
        if learning_config['setup_chosen']['stmk'] == 'Gleinz':
            pv_data = pd.read_excel(os.path.join(config.grid_data_folder, pv_folder, 'NAP_PV_Gleinz_P.xlsx'))
            q_data = pd.read_excel(os.path.join(config.grid_data_folder, pv_folder, 'NAP_PV_Gleinz_Q.xlsx'))
        else:
            pv_data = pd.read_excel(os.path.join(config.grid_data_folder, pv_folder, 'NAP_PV_Neudau_P.xlsx'))
            q_data = pd.read_excel(os.path.join(config.grid_data_folder, pv_folder, 'NAP_PV_Neudau_Q.xlsx'))
    else:
        loads = pd.read_csv(os.path.join(config.grid_data_folder, folder, 'Load.csv'), sep=';', index_col='id')
        profiles = pd.read_csv(os.path.join(config.grid_data_folder, folder, 'LoadProfile.csv'), sep=';', index_col='time')
        pv_profiles = load_workbook(os.path.join(config.grid_data_folder, folder, r"PV_Profiles.xlsx"))

    if config.detection_methods or config.detection_application:

        if config.detection_application and type(data) is str and data == 'sampled' or type(data) is list:
            variation = setup
            pf.activate_variations('Training data Setup')     #activate both PVs here

            """if 'Test Setup A' not in [var.loc_name for var in pf.app.GetActiveNetworkVariations()]:
                pf.activate_variations('Test Setup A')     #activate both PVs here
            if 'Test Setup B' not in [var.loc_name for var in pf.app.GetActiveNetworkVariations()]:
                pf.activate_variations('Test Setup B')"""
        elif type(data) is pd.DataFrame or extract_profiles:
            variation = setup
        elif sim_setting in ['stmk']:
            variation = 'add_chars'
        else:
            variation = learning_config['setup_chosen'].split('_')[1]

        if variation is not None and data is str and type(data) is not list and sim_setting not in ['stmk']:
            pf.activate_variations('Test Setup ' + variation)
            if variation == 'A':
                variation = 'B'
            else:
                variation = 'A'
            pf.deactivate_variations('Test Setup ' + variation)

        chars_dict[element] = {}

        if config.detection_application and not extract_profiles:

            if type(data) is str and data == 'sampled':
                num_data_points = config.num_training_samples

                if element_type == 'PV':
                    if element.loc_name.split(' ')[-1] == setup:
                        mu = config.load_estimation_training_data_distributions_dict['PV']['mu']
                        sigma = config.load_estimation_training_data_distributions_dict['PV']['sigma']
                        if config.training_data_dist == 'standard':
                            samples = np.random.default_rng().normal(mu, sigma, (1, num_data_points))
                        else:
                            samples = np.random.uniform(low=0.0, high=6.0, size=(1, num_data_points))
                        p_profile = samples[0]
                    else:
                        p_profile = np.zeros(num_data_points)     #PV deactivated here

                else:
                    mu = config.load_estimation_training_data_distributions_dict[sim_setting][element.loc_name][0]
                    sigma = config.load_estimation_training_data_distributions_dict[sim_setting][element.loc_name][1]
                    q_factor = config.load_estimation_training_data_distributions_dict[sim_setting][element.loc_name][2]
                    if config.training_data_dist == 'standard':
                        samples_p = np.random.default_rng().normal(mu, sigma, (1, num_data_points))
                        samples_q = np.random.default_rng().normal(mu/q_factor, sigma/q_factor, (1, num_data_points))
                    else:
                        high = mu+3*sigma
                        low = high * config.load_estimation_training_data_distributions_dict[sim_setting][element.loc_name][3]
                        p_low = config.load_estimation_training_data_distributions_dict[sim_setting][element.loc_name][4]
                        samples_p = np.random.uniform(low=p_low, high=high, size=(1, num_data_points))
                        samples_q = np.random.uniform(low=low/q_factor, high=high/q_factor, size=(1, num_data_points))
                    p_profile = samples_p[0]
                    q_profile = samples_q[0]

                begin = '01.01.2022 00:00'
                t_start = pd.Timestamp(begin, tz='utc')
                times = pd.date_range(begin, periods=num_data_points, freq="S")
                t_end = pd.Timestamp(times[-1], tz='utc')

                # Objects which control the time scale
                utc = datetime.timezone.utc
                times = pd.date_range(start=t_start, end=t_end, freq='S',
                                      tz='utc')

                # scale for cyprus profiles (5 minute resolution):
                o_TriTime = pf.create_time_scale(
                    time_scale=f'timescale_training',
                    time_points=times,
                    unit="Y",
                    parent=None,
                    destination_timezone=utc
                )

                p_char = pf.create_vector_characteristic(
                    characteristic=f'p_{element.loc_name}',
                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                    # i].values / factor, index = config.times_household),
                    vector_nodes=p_profile,
                    scale=o_TriTime,
                    usage=2,  # 0,1,2 ... 2 means absolute
                    approximation="constant",
                    parent=None
                )

                chars_dict[element][f'p'] = p_char
                chars_dict[element][f't_start'] = t_start
                chars_dict[element][f't_end'] = t_end

                if element_type == 'load':
                    q_char = pf.create_vector_characteristic(
                        characteristic=f'q_{element.loc_name}',
                        # adds first letter of column name to characteristics name > most commonly P, Q or V
                        # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                        # i].values / factor, index = config.times_household),
                        vector_nodes=q_profile,
                        scale=o_TriTime,
                        usage=2,  # 0,1,2 ... 2 means absolute
                        approximation="constant",
                        parent=None
                    )

                    chars_dict[element][f'q'] = q_char

            elif type(data) is list:

                num_data_points = len(list(profiles[0][f'PV {setup}_P'].values))

                if element_type == 'PV':
                    if element.loc_name.split(' ')[-1] == setup:
                        p_profile = list(profiles[0][f'PV {setup}_P'].values)
                    else:
                        p_profile = np.zeros(num_data_points)     #PV deactivated here

                else:
                    p_load = profiles[1][f'{element.loc_name}_P']
                    q_load = profiles[1][f'{element.loc_name}_Q']

                    #p_measurement_test_bay = profiles[0]['Test Bay ' + config.load_test_bays_map_dict[sim_setting][element.loc_name] + '_p']/1000000
                    #q_measurement_test_bay = profiles[0]['Test Bay ' + config.load_test_bays_map_dict[sim_setting][element.loc_name] + '_q']/1000000

                    """if config.load_test_bays_map_dict[sim_setting][element.loc_name] == config.PV_test_bays_map_dict[sim_setting][setup]:
                        pv_p_profile = profiles[0][f'PV {setup}_P']
                        pv_q_profile = profiles[0][f'PV {setup}_Q']"""

                    p_profile = list(p_load.values)
                    q_profile = list(q_load.values)

                begin = '01.01.2022 00:00'
                t_start = pd.Timestamp(begin, tz='utc')
                times = pd.date_range(begin, periods=num_data_points, freq="S")
                t_end = pd.Timestamp(times[-1], tz='utc')

                # Objects which control the time scale
                utc = datetime.timezone.utc
                times = pd.date_range(start=t_start, end=t_end, freq='S',
                                      tz='utc')

                # scale for cyprus profiles (5 minute resolution):
                o_TriTime = pf.create_time_scale(
                    time_scale=f'timescale_training',
                    time_points=times,
                    unit="Y",
                    parent=None,
                    destination_timezone=utc
                )

                p_char = pf.create_vector_characteristic(
                    characteristic=f'p_{element.loc_name}',
                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                    # i].values / factor, index = config.times_household),
                    vector_nodes=p_profile,
                    scale=o_TriTime,
                    usage=2,  # 0,1,2 ... 2 means absolute
                    approximation="constant",
                    parent=None
                )

                chars_dict[element][f'p'] = p_char
                chars_dict[element][f't_start'] = t_start
                chars_dict[element][f't_end'] = t_end

                if element_type == 'load':
                    q_char = pf.create_vector_characteristic(
                        characteristic=f'q_{element.loc_name}',
                        # adds first letter of column name to characteristics name > most commonly P, Q or V
                        # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                        # i].values / factor, index = config.times_household),
                        vector_nodes=q_profile,
                        scale=o_TriTime,
                        usage=2,  # 0,1,2 ... 2 means absolute
                        approximation="constant",
                        parent=None
                    )

                    chars_dict[element][f'q'] = q_char

            else:

                if config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']:
                    if element_type == 'load':
                        profiles_dict = load_script.finalize_profiles()
                    else:
                        pv_profiles_dict = load_script.create_PV_profiles(save_profile=False)

                num_data_points = 60 - 36  # 9am to 3pm
                for sample_no in list(range(len(data.index)%num_data_points)):
                    start_indices = sample_no % (num_data_points)
                    element_name = element.loc_name
                    if element_type == 'load':
                        p_profile = profiles[element_name + '_P'][start_indices:start_indices+num_data_points+1]  #SHOULD BE IN KW!
                        q_profile = profiles[element_name + '_Q'][start_indices:start_indices+num_data_points+1]
                    else:
                        p_profile = pv_input[start_indices:start_indices+num_data_points+1]


                    begin = f'{sample_no+1}.01.2022 09:00'
                    end = f'{sample_no+1}.01.2022 15:00'    #check if 15:00 or 15:15 to get all datapoints

                    t_start = pd.Timestamp(begin, tz='utc')
                    t_end = pd.Timestamp(end, tz='utc')

                    # Objects which control the time scale
                    utc = datetime.timezone.utc
                    times = pd.date_range(start=t_start, end=t_end, freq=config.resolution,
                                          tz='utc')

                    # scale for cyprus profiles (5 minute resolution):
                    o_TriTime = pf.create_time_scale(
                        time_scale=f'times_{sample_no}',
                        time_points=times,
                        unit="Y",
                        parent=None,
                        destination_timezone=utc
                    )

                    if element_type == 'load':
                        data = data.round()

                    if (config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2',
                                                                                   'phase2']) and element_type == 'load':
                        name = ''.join(element.loc_name.split(' '))
                        if name != 'LB3':
                            data_p_no_DSM = profiles_dict[''.join(element.loc_name.split(' ')) + '_no_DSM']['p'][str(sample_no+1)]
                            data_q_no_DSM = profiles_dict[''.join(element.loc_name.split(' ')) + '_no_DSM']['q'][str(sample_no+1)]
                            data_p_DSM = \
                                profiles_dict[''.join(element.loc_name.split(' ')) + '_DSM']['p'][str(sample_no+1)]
                            data_q_DSM = \
                                profiles_dict[''.join(element.loc_name.split(' ')) + '_DSM']['q'][str(sample_no+1)]

                            p_char = pf.create_vector_characteristic(
                                characteristic=f'p_{element.loc_name}_{sample_no+1}_no_DSM',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=data_p_no_DSM,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'p_{sample_no+1}_noDSM'] = p_char
                            chars_dict[element][f'{sample_no+1}_noDSM_t_start'] = t_start
                            chars_dict[element][f'{sample_no+1}_noDSM_t_end'] = t_end

                            q_char = pf.create_vector_characteristic(
                                characteristic=f'q_{element.loc_name}_{sample_no+1}_no_DSM',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=data_q_no_DSM,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'q_{sample_no+1}_noDSM'] = q_char

                            p_char = pf.create_vector_characteristic(
                                characteristic=f'p_{element.loc_name}_{sample_no+1}_DSM',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=data_p_DSM,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'p_{sample_no+1}_DSM'] = p_char
                            chars_dict[element][f'{sample_no+1}_DSM_t_start'] = t_start
                            chars_dict[element][f'{sample_no+1}_DSM_t_end'] = t_end

                            q_char = pf.create_vector_characteristic(
                                characteristic=f'q_{element.loc_name}_{sample_no+1}_DSM',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=data_q_DSM,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'q_{sample_no+1}_DSM'] = q_char

                        else:
                            p_profile = profiles_dict[''.join(element.loc_name.split(' '))]['p'][str(sample_no+1)]
                            q_profile = profiles_dict[''.join(element.loc_name.split(' '))]['q'][str(sample_no+1)]

                            p_char = pf.create_vector_characteristic(
                                characteristic=f'p_{element.loc_name}_{sample_no+1}',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=p_profile,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'p_{sample_no+1}'] = p_char
                            chars_dict[element][f't_{sample_no+1}_start'] = t_start
                            chars_dict[element][f't_{sample_no+1}_end'] = t_end

                            q_char = pf.create_vector_characteristic(
                                characteristic=f'q_{element.loc_name}_{sample_no+1}',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=q_profile,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'q_{sample_no+1}'] = q_char

                    else:
                        if config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']: p_profile = \
                        pv_profiles_dict[str(sample_no+1)]['absolute P (10kWp)']

                        p_char = pf.create_vector_characteristic(
                            characteristic=f'p_{element.loc_name}_{sample_no+1}',
                            # adds first letter of column name to characteristics name > most commonly P, Q or V
                            # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                            # i].values / factor, index = config.times_household),
                            vector_nodes=p_profile,
                            scale=o_TriTime,
                            usage=2,  # 0,1,2 ... 2 means absolute
                            approximation="constant",
                            parent=None
                        )

                        chars_dict[element][f'p_{sample_no+1}'] = p_char
                        chars_dict[element][f't_{sample_no+1}_start'] = t_start
                        chars_dict[element][f't_{sample_no+1}_end'] = t_end

                        if element_type == 'load':
                            q_char = pf.create_vector_characteristic(
                                characteristic=f'q_{element.loc_name}_{sample_no+1}',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=q_profile,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'q_{sample_no+1}'] = q_char

        elif sim_setting in ['stmk']:
            v_profiles_dict = {}
            pv_profiles_dict = {}
            q_profiles_dict = {}
            steps_per_day = 1440

            if learning_config['setup_chosen']['stmk'] == 'Gleinz':
                voltage_data = voltage_data.loc[11 * steps_per_day:, :].reset_index(drop=True) #cut away first 11 days as PV doesnt do anything here
                pv_data = pv_data.loc[11 * steps_per_day:, :].reset_index(drop=True)
                q_data = q_data.loc[11 * steps_per_day:, :].reset_index(drop=True)

            else:
                voltage_data = voltage_data.loc[0 * steps_per_day:, :].reset_index(
                    drop=True)  # cut away first 0 days as PV doesnt do anything here
                pv_data = pv_data.loc[0 * steps_per_day:, :].reset_index(drop=True)
                q_data = q_data.loc[0 * steps_per_day:, :].reset_index(drop=True)

            #slice 1440 stepes for one day
            for i in range((int(len(voltage_data) / steps_per_day))):  # This ensures all rows are captured
                v_profiles_dict[str(i)] = voltage_data.loc[i * steps_per_day:(i + 1) * steps_per_day, :][:-1]   #1440-1 since 0 indexed
                pv_profiles_dict[str(i)] = pv_data.loc[i * steps_per_day:(i + 1) * steps_per_day, :][:-1]       #1440-1 since 0 indexed
                q_profiles_dict[str(i)] = q_data.loc[i * steps_per_day:(i + 1) * steps_per_day, :][:-1]

            for day in v_profiles_dict:
                if config.dev_mode and day == '2': break
                begin = v_profiles_dict[day].loc[v_profiles_dict[day].index[0], 'Utc']
                try:
                    end = v_profiles_dict[day].loc[(steps_per_day) * (int(day) + 1) - 1, 'Utc']
                except KeyError:
                    continue


                t_start = pd.Timestamp(begin, tz='utc')
                t_start_unix = convert_to_unix_timestamp(t_start)
                t_end = pd.Timestamp(end, tz='utc')
                t_end_unix = convert_to_unix_timestamp(t_end)

                # Objects which control the time scale
                utc = datetime.timezone.utc
                times = pd.date_range(start=t_start, end=t_end, freq=config.resolution,
                                      tz='utc')

                # scale for netze steiermark profiles (1 minute resolution):
                o_TriTime = pf.create_time_scale(
                    time_scale=f'times_{day}',
                    time_points=times,
                    unit="Y",
                    parent=None,
                    destination_timezone=utc
                )

                if element_type == 'PV' or element_type == 'Genstat':
                    data = pv_profiles_dict[day]

                    p_char = pf.create_vector_characteristic(
                        characteristic=f'p_{element.loc_name}_{day}',
                        # adds first letter of column name to characteristics name > most commonly P, Q or V
                        # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                        # i].values / factor, index = config.times_household),
                        vector_nodes=data[data.columns[2]].values*1000,
                        scale=o_TriTime,
                        usage=2,  # 0,1,2 ... 2 means absolute
                        approximation="constant",
                        parent=None
                    )

                    chars_dict[element][f'p_{day}'] = p_char
                    chars_dict[element][f'{day}_t_start'] = t_start
                    chars_dict[element][f'{day}_t_end'] = t_end

                    if element_type == 'Genstat' or element.loc_name == 'PV_System_as_is':
                        data = q_profiles_dict[day]

                        q_char = pf.create_vector_characteristic(
                            characteristic=f'q_{element.loc_name}_{day}',
                            # adds first letter of column name to characteristics name > most commonly P, Q or V
                            # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                            # i].values / factor, index = config.times_household),
                            vector_nodes=data[data.columns[2]].values*(1)*1000,
                            scale=o_TriTime,
                            usage=2,  # 0,1,2 ... 2 means absolute
                            approximation="constant",
                            parent=None
                        )

                        chars_dict[element][f'q_{day}'] = q_char
                        chars_dict[element][f'{day}_t_start'] = t_start
                        chars_dict[element][f'{day}_t_end'] = t_end

                elif element_type == 'Xnet':
                    data = v_profiles_dict[day]

                    v_char = pf.create_vector_characteristic(
                        characteristic=f'p_{element.loc_name}_{day}',
                        # adds first letter of column name to characteristics name > most commonly P, Q or V
                        # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                        # i].values / factor, index = config.times_household),
                        vector_nodes=v_profiles_dict[day]['mean voltage p.u.'].values,
                        scale=o_TriTime,
                        usage=2,  # 0,1,2 ... 2 means absolute
                        approximation="constant",
                        parent=None
                    )

                    chars_dict[element][f'v_{day}'] = v_char
                    chars_dict[element][f'{day}_t_start'] = t_start
                    chars_dict[element][f'{day}_t_end'] = t_end


        else:
            
            if config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']:
                if element_type == 'load':
                    profiles_dict = load_script.finalize_profiles()
                else:
                    pv_profiles_dict = load_script.create_PV_profiles(save_profile=False)

            for id, row in loads.iterrows():

                if element.loc_name == 'LB 2':
                    if id == 'LV4.101 Load 11': break
                else:
                    if id == 'LV4.101 Load 30': break

                p_profile = profiles[row['profile'] + '_pload']
                q_profile = profiles[row['profile'] + '_qload']
                p_load = row['pLoad']
                q_load = row['qLoad']

                i = 1
                data = pd.DataFrame()

                # for sheet_name in pv_profiles.sheetnames[1:]:
                profiles_and_scalings = [(['Tabelle' + i for i in ['22', '9', '14', '32', '12', '2', '8']], 10),
                                         (['Tabelle' + i for i in ['15', '23', '24', '30', '27', '3', '6', '10']], 20)]
                for item in profiles_and_scalings:
                    for sheet_name in item[0]:
                        sheet_df = pd.read_excel(os.path.join(config.grid_data_folder, folder, "PV_Profiles.xlsx"), sheet_name=sheet_name, header=None)

                        #begin = sheet_df[0][0].strftime("%d.%m.%Y %H:%M")
                        #end = sheet_df[0][95].strftime("%d.%m.%Y %H:%M")

                        start_index = 36  # 9am
                        end_index = 60  # 3pm

                        begin = sheet_df[0][start_index].strftime("%d.%m.%Y %H:%M")
                        end = sheet_df[0][end_index].strftime("%d.%m.%Y %H:%M")

                        data = sheet_df[2][ start_index:end_index+1]

                        if element_type == 'load':

                            p_slice = p_profile[begin:end]
                            p_values = p_slice * p_load * 1000  # to have values in kW

                            q_slice = q_profile[begin:end]
                            q_values = q_slice * q_load * 1000  # to have values in kW

                            # powerfactors = [math.cos(math.atan(i)) for i in q_values / p_values] not necessary here

                            if element.loc_name == 'LB 7 8':
                                factor = 2
                            else:
                                factor = 1
                            p_values = p_values * item[1] * factor
                            q_values = q_values * item[1] * factor

                            data = p_values

                        # string_t_start = '2017-01-01 00:00:00'
                        # t_start = pd.Timestamp(string_t_start, tz='utc')
                        t_start = pd.Timestamp(begin, tz='utc')
                        #t_start = pd.Timestamp('2017-01-01 00:00:00', tz='utc')
                        t_start_unix = convert_to_unix_timestamp(t_start)
                        t_end = pd.Timestamp(end, tz='utc')
                        t_end_unix = convert_to_unix_timestamp(t_end)
                        #t_end = pd.Timestamp('2018-01-01 00:00:00', tz='utc') - pd.Timedelta(config.resolution)

                        # Objects which control the time scale
                        utc = datetime.timezone.utc
                        times = pd.date_range(start=t_start, end=t_end, freq=config.resolution,
                                              tz='utc')

                        # scale for cyprus profiles (5 minute resolution):
                        o_TriTime = pf.create_time_scale(
                            time_scale=f'times_{sheet_name}',
                            time_points=times,
                            unit="Y",
                            parent=None,
                            destination_timezone=utc
                        )

                        if element_type == 'load':
                            data = data.round()

                        if (config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']) and element_type == 'load':
                            name = ''.join(element.loc_name.split(' '))
                            if name != 'LB3':
                                data_p_no_DSM = profiles_dict[''.join(element.loc_name.split(' ')) + '_no_DSM']['p'][str(i)]
                                data_q_no_DSM = profiles_dict[''.join(element.loc_name.split(' ')) + '_no_DSM']['q'][str(i)]
                                data_p_DSM = \
                                profiles_dict[''.join(element.loc_name.split(' ')) + '_DSM']['p'][str(i)]
                                data_q_DSM = \
                                profiles_dict[''.join(element.loc_name.split(' ')) + '_DSM']['q'][str(i)]

                                p_char = pf.create_vector_characteristic(
                                    characteristic=f'p_{element.loc_name}_{sheet_name}_no_DSM',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_p_no_DSM,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'p_{i}_noDSM'] = p_char
                                chars_dict[element][f'{i}_noDSM_t_start'] = t_start
                                chars_dict[element][f'{i}_noDSM_t_end'] = t_end

                                q_char = pf.create_vector_characteristic(
                                    characteristic=f'q_{element.loc_name}_{sheet_name}_no_DSM',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_q_no_DSM,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'q_{i}_noDSM'] = q_char

                                p_char = pf.create_vector_characteristic(
                                    characteristic=f'p_{element.loc_name}_{sheet_name}_DSM',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_p_DSM,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'p_{i}_DSM'] = p_char
                                chars_dict[element][f'{i}_DSM_t_start'] = t_start
                                chars_dict[element][f'{i}_DSM_t_end'] = t_end

                                q_char = pf.create_vector_characteristic(
                                    characteristic=f'q_{element.loc_name}_{sheet_name}_DSM',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_q_DSM,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'q_{i}_DSM'] = q_char

                            else:
                                data_p = profiles_dict[''.join(element.loc_name.split(' '))]['p'][str(i)]
                                data_q = profiles_dict[''.join(element.loc_name.split(' '))]['q'][str(i)]

                                p_char = pf.create_vector_characteristic(
                                    characteristic=f'p_{element.loc_name}_{sheet_name}',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_p,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'p_{i}'] = p_char
                                chars_dict[element][f'{i}_t_start'] = t_start
                                chars_dict[element][f'{i}_t_end'] = t_end

                                q_char = pf.create_vector_characteristic(
                                    characteristic=f'q_{element.loc_name}_{sheet_name}',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=data_q,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'q_{i}'] = q_char

                        else:
                            if config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']: data = pv_profiles_dict[str(i)]['absolute P (10kWp)']
                            p_char = pf.create_vector_characteristic(
                                characteristic=f'p_{element.loc_name}_{sheet_name}',
                                # adds first letter of column name to characteristics name > most commonly P, Q or V
                                # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                # i].values / factor, index = config.times_household),
                                vector_nodes=data,
                                scale=o_TriTime,
                                usage=2,  # 0,1,2 ... 2 means absolute
                                approximation="constant",
                                parent=None
                            )

                            chars_dict[element][f'p_{i}'] = p_char
                            chars_dict[element][f'{i}_t_start'] = t_start
                            chars_dict[element][f'{i}_t_end'] = t_end

                            if element_type == 'load':
                                q_char = pf.create_vector_characteristic(
                                    characteristic=f'q_{element.loc_name}_{sheet_name}',
                                    # adds first letter of column name to characteristics name > most commonly P, Q or V
                                    # vector_nodes= pd.Series(pd.read_csv('output/' + profile, sep=';', decimal='.', index_col=0, header=0).iloc[:,
                                    # i].values / factor, index = config.times_household),
                                    vector_nodes=q_values,
                                    scale=o_TriTime,
                                    usage=2,  # 0,1,2 ... 2 means absolute
                                    approximation="constant",
                                    parent=None
                                )

                                chars_dict[element][f'q_{i}'] = q_char

                        i += 1
                if config.sim_setting == 'ERIGrid_phase_2' or sim_setting in ['ERIGrid_phase_2', 'phase2']:
                    break

    return chars_dict

"""def load_profiles():


    profiles = {}
    profiles = create_characteristics(o_ElmLod, profiles, sim_setting=config.sim_setting,
                                    type='load', data=data, setup=setup,
                                    extract_profiles=extract_profiles)  # create characteristics

    profiles = create_characteristics(o_PV, profiles, sim_setting=config.sim_setting, type='PV', data=data,
                                    setup=setup, extract_profiles=extract_profiles)  # create charcteristics
    return profiles"""


def prepare_grid(app, file, o_ElmNet, data=None, setup=None, phase=None, extract_profiles=False, pv_input=None, estimation=None):
    # set path for load and generation profiles
    char_folder = app.GetProjectFolder('chars')
    if config.deeplearning:
        chars = list(
            pd.read_csv(os.path.join(config.grid_data_folder, file, 'LoadProfile.csv'), sep=';',
                        index_col='time').columns) \
                + list(
            pd.read_csv(os.path.join(config.grid_data_folder, file, 'RESProfile.csv'), sep=';',
                        index_col='time').columns)
        for char_name in chars:
            char = char_folder.SearchObject(char_name + '.ChaTime')
            if os.name == 'nt':
                init_f_name_ending = char.f_name.split('\\')[-1]
            else:
                init_f_name_ending = char.f_name.split('/')[-1]
            char.f_name = os.path.join(config.grid_data_folder, file, init_f_name_ending)

    # loads_df = pd.read_csv(os.path.join(config.grid_data_folder, file, 'Load.csv'), sep=';')
    loads_by_type = {'regular_loads': [], 'heatpumps': [], 'EV_charging_stations': {'Work': [], 'Home': []}}

    if config.detection_methods or config.detection_application:
        curves = {}
        EV_charging_stations = None

    for o_ElmLod in app.GetCalcRelevantObjects('.ElmLod'):  # has to be done like this to activated profiles
        if config.deeplearning:
            o_ChaTime_p = pf.get_referenced_characteristics(o_ElmLod, 'plini')[
                0]  # get P characteristic (profile) from load
            o_ChaTime_q = pf.get_referenced_characteristics(o_ElmLod, 'qlini')[  # same for Q (reactive Power)
                0]

            if o_ChaTime_p.loc_name[0] in ['H', 'G'] and o_ChaTime_p.loc_name[1].isnumeric():
                loads_by_type['regular_loads'].append((o_ElmLod, o_ChaTime_p, o_ChaTime_q))
            elif o_ChaTime_p.loc_name.split('_')[0] in ['Air', 'Soil']:
                loads_by_type['heatpumps'].append((o_ChaTime_p, o_ChaTime_q))
                o_ElmLod.Delete()  # delete to make space for own setup
            elif o_ChaTime_p.loc_name.split('_')[0] in ['HLS', 'APLS']:
                if o_ChaTime_p.loc_name.split('_')[0] == 'HLS':
                    loads_by_type['EV_charging_stations']['Home'].append(
                        ((o_ChaTime_p, o_ElmLod.plini), (o_ChaTime_q, o_ElmLod.qlini), o_ElmLod.typ_id))
                else:
                    loads_by_type['EV_charging_stations']['Work'].append(
                        ((o_ChaTime_p, o_ElmLod.plini), (o_ChaTime_q, o_ElmLod.qlini), o_ElmLod.typ_id))

                o_ElmLod.Delete()  # delete to make space for own setup
            else:
                print('Unknown load type found!')

            pf.set_referenced_characteristics(o_ElmLod, 'plini', o_ChaTime_p)  # set characteristic for load

            pf.set_referenced_characteristics(o_ElmLod, 'qlini', o_ChaTime_q)

        if config.detection_methods:
            curves = create_characteristics(o_ElmLod, curves, sim_setting=config.sim_setting,
                                            element_type='load')  # create charcteristics

        if config.detection_application:
            if type(data) is str and data == 'sampled':

                curves = create_characteristics(o_ElmLod, curves, sim_setting=phase.split('_')[-1],
                                                element_type='load', data=data, setup=setup)  # create characteristics
            elif extract_profiles:
                curves = create_characteristics(o_ElmLod, curves, sim_setting=phase.split('_')[-1],
                                                element_type='load', data=data, setup=setup, extract_profiles=extract_profiles)  # create characteristics
            elif type(data) is list:
                curves = create_characteristics(o_ElmLod, curves, sim_setting=phase.split('_')[-1],
                                                element_type='load', data=data, setup=setup)  # create characteristics
            elif estimation:
                curves = create_characteristics(o_ElmLod, curves, sim_setting=phase.split('_')[-1],
                                                element_type='load', data=data, setup=setup)  # create characteristics
            else:
                curves = create_characteristics(o_ElmLod, curves, sim_setting=config.sim_setting,
                                                element_type='load', data=data, setup=setup)  # create characteristics

    if config.detection_methods:
        for o_PV in app.GetCalcRelevantObjects('.ElmPvsys'):
            curves = create_characteristics(o_PV, curves, sim_setting=config.sim_setting, element_type='PV')
        for o_PV in app.GetCalcRelevantObjects('.ElmGenstat'):
            curves = create_characteristics(o_PV, curves, sim_setting=config.sim_setting, element_type='Genstat')
        if config.sim_setting in ['stmk']:
            for o_Xnet in app.GetCalcRelevantObjects('.ElmXnet'):
                curves = create_characteristics(o_Xnet, curves, sim_setting=config.sim_setting, element_type='Xnet')


    if config.detection_application:
        if type(data) == str and data == 'sampled':
            for o_PV in app.GetCalcRelevantObjects('.ElmPvsys'):
                curves = create_characteristics(o_PV, curves, sim_setting=config.sim_setting, element_type='PV', data=data,
                                                setup=setup)        # create charcteristics
        elif extract_profiles:
            for o_PV in app.GetCalcRelevantObjects('.ElmPvsys'):
                curves = create_characteristics(o_PV, curves, sim_setting=config.sim_setting, element_type='PV', data=data,
                                                setup=setup, extract_profiles=extract_profiles)        # create charcteristics
        else:
            for o_PV in app.GetCalcRelevantObjects('.ElmPvsys'):
                if estimation:
                    curves = create_characteristics(o_PV, curves, sim_setting=phase.split('_')[-1],
                                                element_type='PV', data=data, setup=setup, pv_input=pv_input)  # create characteristics
                else:
                    curves = create_characteristics(o_PV, curves, sim_setting=config.sim_setting, element_type='PV', data=data, setup=setup, pv_input=pv_input)  # create charcteristics

    if config.deeplearning:
        # deactivate storages in grid and count PVs for later use
        for o_ElmGenstat in app.GetCalcRelevantObjects('.ElmGenstat'):
            if o_ElmGenstat.cCategory in ['Storage', 'Batterie'] and config.percentage['BESS'] == 0:
                o_ElmGenstat.Delete()  # first copy properties? then delete BESS in order to make space for own setup
            elif o_ElmGenstat.cCategory in ['Photovoltaic', 'Fotovoltaik']:
                o_ChaTime = pf.get_referenced_characteristics(o_ElmGenstat, 'pgini')[
                    0]  # get characteristic (profile) from original PV
                # PV_apparent_power = o_ElmGenstat.sgn
                o_ElmGenstat.Delete()  # delete PV in order to make space for own setup; PVs placed are scaled with adjacent load
                if len(pf.get_referenced_characteristics(o_ElmGenstat, 'pgini')) > 1:
                    print(
                        'More than one PV profile found; consider which one to choose (default: first one found chosen)')

        curves = place_PVs(app, o_ElmNet, o_ChaTime, loads_by_type,
                           PV_apparent_power=0.005)  # PV_apparent_power=0.005 means 5kWp
        EV_charging_stations = place_EVCS(o_ElmNet, loads_by_type)

    utf8_complaint_naming(o_ElmNet)  # check if element names are UTF8 compliant and rename if not

    return (curves, EV_charging_stations, loads_by_type['regular_loads'])
