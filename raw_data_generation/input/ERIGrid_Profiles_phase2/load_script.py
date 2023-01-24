import pandas as pd
import math
from datetime import datetime
import os
import operator

from openpyxl import load_workbook
import xlsxwriter

def check_maximum_load(max=100, adjusted=False, print_updates=False):

    flag = 0
    count = 0
    reduce_dict = {}
    combinations = [['1_no_DSM(use_always_in_Setup_A)', '2_DSM', '3_(use_in_all_Setups)'],
                    ['1_no_DSM(use_always_in_Setup_A)', '2_no_DSM(use_always_in_Setup_B)',
                     '3_(use_in_all_Setups)'], ['1_DSM', '2_no_DSM(use_always_in_Setup_B)', '3_(use_in_all_Setups)']]

    for scen in list(range(1,16)):
        for combination in combinations:
            for file in combination:
                reduce_dict['ERI-Grid - Scenario ' + str(scen) + '.' + file + '.txt'] = {}

    for scen in list(range(1,16)):
        for combination in combinations:
            lines_of_file = {}
            for file in combination:
                try:
                    with open('ERI-Grid - Scenario ' + str(scen) + '.' + file + '.txt') as f:
                        lines_of_file[file] = f.readlines()
                except FileNotFoundError:
                    with open(
                            os.path.join(profiles_path, 'ERI-Grid - Scenario ' + str(scen) + '.' + file + '.txt')) as f:
                        lines_of_file[file] = f.readlines()
            step_count = 0
            for line_number in list(range(0,len(lines_of_file[combination[0]]))):
                appearant_power_sum = 0
                for file in combination:
                    line = lines_of_file[file][line_number]
                    if line.split(' ')[0] == 'Load':
                        step_count +=1
                        real_power = int(line.split(' ')[1][:-2])
                        power_factor = float(line.split(' ')[2][:-4])
                        if adjusted and file == '3_(use_in_all_Setups)':
                            appearant_power = 4.4   # maximum value of ERI-Grid - Scenario manual.3_(use_in_all_Setups): to check if profile for manual laod control device is also not to high
                        else:
                            appearant_power = real_power / power_factor                      #calculates S at certain point in time for certain load
                        appearant_power_sum = appearant_power_sum + appearant_power
                if appearant_power_sum > max:
                    flag = 1
                    count += 1
                    reduction_factor = max / appearant_power_sum
                    for file in combination:
                        reduce_dict['ERI-Grid - Scenario ' + str(scen) + '.' + file +  '.txt'][int(step_count/3)] = reduction_factor
                    if print_updates:
                        print(f'Sum of loads is over the maximum for scenario {scen} in combination {combination} at step {int(step_count/3)} with an appearant power of {appearant_power_sum}: load has to be reduced by a factor of {reduction_factor}')


    return flag, count, reduce_dict

def create_profile(setting, row, reduce_dict={}, file_type='txt', save_profiles=True):

    if file_type == 'excel':
        workbook = xlsxwriter.Workbook('Load Profiles.xlsx')

    if setting.split('_')[0] == 'LB2':
        factor = 2
        addition = '(use_always_in_Setup_B)'
    if setting.split('_')[0] in ['LB1', 'LB3']:
        factor = 1
        addition = '(use_always_in_Setup_A)'
    try:
        if setting.split('_')[1] == 'no':
            dsm = 'no_DSM' + addition
        if setting.split('_')[1] == 'DSM':
            dsm = 'DSM'
    except IndexError:
        dsm = '(use_in_all_Setups)'

    p_profile = profiles[row['profile'] + '_pload']
    q_profile = profiles[row['profile'] + '_qload']
    p_load = row['pLoad']
    q_load = row['qLoad']
    p_values_dict = {}
    q_values_dict = {}

    i = 1
    data = pd.DataFrame()

    # for sheet_name in pv_profiles.sheetnames[1:]:
    profiles_and_scalings = [(['Tabelle' + i for i in ['22', '9', '14', '32', '12', '2', '8']], 10),
                             (['Tabelle' + i for i in ['15', '23', '24', '30', '27', '3', '6', '10']], 20)]
    for item in profiles_and_scalings:
        for sheet_name in item[0]:
            try:
                profiles_path = os.path.join('raw_data_generation', 'input', 'ERIGrid_Profiles_phase2')
                sheet_df = pd.read_excel(os.path.join(profiles_path, 'PV_Profiles.xlsx'), sheet_name=sheet_name, header=None)
            except FileNotFoundError:
                sheet_df = pd.read_excel('PV_Profiles.xlsx', sheet_name=sheet_name,
                                         header=None)

            begin = sheet_df[0][0].strftime("%d.%m.%Y %H:%M")
            end = sheet_df[0][95].strftime("%d.%m.%Y %H:%M")

            p_slice = p_profile[begin:end]
            p_values = p_slice * p_load * 1000

            start_index = 36  # 9am
            end_index = 61  # 3pm

            q_slice = q_profile[begin:end]
            q_values = q_slice * q_load * 1000

            if dsm == 'DSM':  # find slice with highest consumption in data and then use this slice as laod as to perform DSM by shifting the highest laod where there is PV
                windows = [i for i in p_values.rolling(end_index - start_index)]
                consumption_in_window = [i.sum() for i in windows]
                index, value = max(enumerate(consumption_in_window), key=operator.itemgetter(1))
                window_with_highest_load = windows[index]

                q_windows = [i for i in q_values.rolling(end_index - start_index)]
                q_window_with_highest_load = q_windows[index]

                p_shifted = window_with_highest_load
                q_shifted = q_window_with_highest_load

                #for plotting without any maximum check
                #pv_and_load = pd.DataFrame(data={'pv': sheet_df[2].values[start_index:end_index], 'load_regular': p_values.values[start_index:end_index], 'load_shifted': p_shifted.values}, columns=['pv', 'load_regular', 'load_shifted'], index =sheet_df[0][start_index:end_index].values)
                #pv_and_load.plot()

                p_values = p_shifted
                q_values = q_shifted

            powerfactors = [math.cos(math.atan(i)) for i in q_values / p_values]

            if file_type == 'excel':
                worksheet = workbook.add_worksheet(sheet_name + '_' + str(counter))

            scores = (
                ["' Project", 'ERI-Grid - Scenario ' + str(i)],
                ['Instrumentation', 'Full'],
                ['Capture', 'On'],
                ['Delay', '00:00:05'],
            )

            if dsm == 'DSM':
                end_index = end_index - start_index
                start_index = 0
            p_values_final = []
            q_values_final = []

            for k in list(range(start_index, end_index)):
                # scores = scores + (['Load ' + f"{format(p_values[k],'.2f')}" + 'kW ' + f"{format(powerfactors[k],'.2f')}" + 'cosØ', '00:00:05'],)
                if dsm == 'DSM':
                    step = k + 1
                else:
                    step = (k-start_index) + 1

                if k - start_index == 0:
                    delay = '00:2:55'
                else:
                    delay = '00:0:55'

                try:
                    entry = reduce_dict['ERI-Grid - Scenario ' + str(i) + '.' + setting.split('_')[0][-1] + '_' + dsm + ".txt"]
                    if type(entry) == tuple:
                        reduction_factor = 1
                        for iteration in entry:
                            if step in iteration:
                                reduction_factor_iteration = float(str(iteration[step])[:4])
                            else:
                                reduction_factor_iteration = 1
                            reduction_factor = reduction_factor * reduction_factor_iteration
                    else:
                        if step in entry:
                            reduction_factor = float(str(entry[step])[:4])
                        else:
                            reduction_factor = 1
                except KeyError:
                    reduction_factor = 1

                scores = scores + (
                    ['Load ' + f"{format(round(p_values[k] * item[1] * factor * reduction_factor), '.0f')}" + 'kW ' + f"{format(powerfactors[k], '.2f')}" + 'cosØ',
                        '00:00:05'],)
                scores = scores + (['Delay', delay],)
                p_values_final.append(p_values[k] * item[1] * factor * reduction_factor)
                q_values_final.append(q_values[k] * item[1] * factor * reduction_factor)

            scores = scores + ([' ', ' '],)
            scores = scores + (['Reject', ' '],)
            scores = scores + (['Capture', 'Off'],)

            # Start from the first cell. Rows and
            # columns are zero indexed.
            row = 0
            col = 0

            p_values_dict[str(i)] = p_values_final
            q_values_dict[str(i)] = q_values_final

            if save_profiles:
                # Iterate over the data and write it out row by row.
                if file_type == 'excel':
                    for name, score in (scores):
                        worksheet.write(row, col, name)
                        worksheet.write(row, col + 1, score)
                        row += 1

                elif file_type == 'txt':
                    # with open('ERI-Grid - Scenario ' + sheet_name.split('e')[2] + '_' + str(counter) + ".txt", "w") as text_file:
                    # with open('ERI-Grid - Scenario ' + sheet_name.split('e')[2] + '_' + str(counter) + ".txt", "w") as text_file:

                    with open(os.path.join(profiles_path, 'ERI-Grid - Scenario ' + str(i) + '.' + setting.split('_')[0][-1] + '_' + dsm + ".txt"),
                              "w") as text_file:

                        for name, score in (scores):
                            print(f"{name} {score}", file=text_file)

                if setting.split('_')[0] == 'LB3' and i == 9:

                    scores = (
                        ["' Project", 'ERI-Grid - Scenario ' + 'manual'],
                        ['Profile for manual application',  'on hand held load control device'],
                    )
                    for k in list(range(start_index, end_index-9)):     #regular sequences are 25 long, for handheld only 16 steps allowed
                        # scores = scores + (['Load ' + f"{format(p_values[k],'.2f')}" + 'kW ' + f"{format(powerfactors[k],'.2f')}" + 'cosØ', '00:00:05'],)
                        if dsm == 'DSM':
                            step = k + 1
                        else:
                            step = (k - start_index) + 1
                        if k - start_index == 0 or step == 16:
                            delay = '00:03:00'
                        elif step in [2, 4, 6, 8, 10, 12, 14]:
                            delay = '00:02:00'
                        else:
                            delay = '00:01:00'

                        try:
                            if step in reduce_dict[
                                'ERI-Grid - Scenario ' + str(i) + '.' + setting.split('_')[0][-1] + '_' + dsm + ".txt"]:
                                reduction_factor = float(str(reduce_dict['ERI-Grid - Scenario ' + str(i) + '.' +
                                                                         setting.split('_')[0][-1] + '_' + dsm + ".txt"][
                                                                 step])[:4])
                            else:
                                reduction_factor = 1
                        except KeyError:
                            reduction_factor = 1

                        scores = scores + (
                            [
                             f"{step} {format(((p_values[k] * item[1] * factor * reduction_factor)/50)*100, '.0f')}%  " + f"{format(powerfactors[k], '.2f')}" + 'cosØ',
                                delay],)

                        # Start from the first cell. Rows and
                        # columns are zero indexed.
                        row = 0
                        col = 0

                        # Iterate over the data and write it out row by row.
                        if file_type == 'excel':
                            for name, score in (scores):
                                worksheet.write(row, col, name)
                                worksheet.write(row, col + 1, score)
                                row += 1

                        elif file_type == 'txt':
                            # with open('ERI-Grid - Scenario ' + sheet_name.split('e')[2] + '_' + str(counter) + ".txt", "w") as text_file:
                            # with open('ERI-Grid - Scenario ' + sheet_name.split('e')[2] + '_' + str(counter) + ".txt", "w") as text_file:
                            with open(os.path.join(profiles_path,
                                                   'ERI-Grid - Scenario ' + 'manual.' + setting.split('_')[0][-1] + '_' + dsm + ".txt"),
                                      "w") as text_file:

                                for name, score in (scores):
                                    print(f"{name} {score}", file=text_file)

            i += 1

    if save_profiles:
        if file_type == 'excel':
            workbook.close()

    return {'p': p_values_dict, 'q': q_values_dict}

def create_profiles(reduce_dict={}, save_profiles=True):

    # load profile is shifted when DSM > window with highest consumption is found and used as load

    final_profiles = {}
    for id, row in loads.iterrows():

        if id == 'LV4.101 Load 1':
            # run for load LB2 (big load): ERI-Grid - Scenario x.2_DSM / _no_DSM(use_always_in_Setup_B)
            final_profiles['LB2_no_DSM'] = create_profile('LB2_no_DSM', row, reduce_dict, save_profiles=save_profiles)
            final_profiles['LB2_DSM'] = create_profile('LB2_DSM', row, reduce_dict, save_profiles=save_profiles)
        if id == 'LV4.101 Load 30':
            # run for load LB1 (small load to be controlled): ERI-Grid - Scenario x.1_DSM / _no_DSM(use_always_in_Setup_A)
            final_profiles['LB1_no_DSM'] = create_profile('LB1_no_DSM', row, reduce_dict, save_profiles=save_profiles)
            final_profiles['LB1_DSM'] = create_profile('LB1_DSM', row, reduce_dict, save_profiles=save_profiles)
        if id == 'LV4.101 Load 17':
            # run for load LB3 (small load NOT to be controlled): ERI-Grid - Scenario x.3 +Plan B (16 steps, only one profile)
            final_profiles['LB3'] = create_profile('LB3', row, reduce_dict, save_profiles=save_profiles)
            # ADD BACKUP WITH 16 STEPS AND LOAD IN % OF RATED LOADBANK (50 kVA) AND PF
        if id == 'LV4.101 Load 18':
            break

    return final_profiles

def merge_dicts(ds):

    d = {}
    for k in ds[0].keys():
        d[k] = tuple(d[k] for d in ds)

    return d

def create_PV_profiles(save_profile = True):

    i = 1
    # for sheet_name in pv_profiles.sheetnames[1:]:
    profiles_and_scalings = [(['Tabelle' + i for i in ['22', '9', '14', '32', '12', '2', '8']], 10),
                             (['Tabelle' + i for i in ['15', '23', '24', '30', '27', '3', '6', '10']], 20)]
    pv_dict = {}
    for item in profiles_and_scalings:
        for sheet_name in item[0]:
            try:
                profiles_path = os.path.join('raw_data_generation', 'input', 'ERIGrid_Profiles_phase2')
                sheet_df = pd.read_excel(os.path.join(profiles_path, 'PV_Profiles.xlsx'), sheet_name=sheet_name,
                                         header=None)
            except FileNotFoundError:
                sheet_df = pd.read_excel('PV_Profiles.xlsx', sheet_name=sheet_name, header=None)

            start_index = 36  # 9am
            end_index = 60  # 3pm

            begin = sheet_df[0][start_index].strftime(("%Y.%m.%d %H:%M:%S"))
            end = sheet_df[0][end_index].strftime(("%Y.%m.%d %H:%M:%S"))

            try:
                profiles = pd.DataFrame(columns=['absolute P (10kWp)'],data=sheet_df['absolute P (10kWp)'].values, index=sheet_df['Date'])
            except KeyError:
                profiles = pd.DataFrame(columns=['absolute P (10kWp)'], data=sheet_df[2].values)

            try:
                p_slice = profiles[begin:end]
            except TypeError:
                p_slice = profiles[start_index:end_index+1]
            i_values = p_slice / 400 * 1000        #400 Volts; P = U*I
            steps = list(range(1,26))
            durations = [180] + [60 for i in steps[1:]]
            contents = pd.DataFrame(columns=['Is(A)', 'Maintain(S)'], data={'Is(A)': i_values['absolute P (10kWp)'].values, 'Maintain(S)': durations})#, index=steps)
            #contents.index.name = 'Step'
            if save_profile:
                contents.to_csv(os.path.join(os.getcwd(), str(i)+ '_PV.csv'), header=True, sep=',', decimal='.', float_format='%.' + '%sf' % 2, index = False)
            pv_dict[str(i)] = p_slice
            i += 1

    return pv_dict

def finalize_profiles(print_updates=False):

    final_profiles = create_profiles()
    flag, count, reduce_dict = check_maximum_load()  # check is maximum load is not higher than certain value (100kVA for phase II)
    if flag == 0:
        if print_updates:
            print('Sum of loads is under the maximum at all times')
    if flag == 1:
        if print_updates:
            print(f'Sum of loads is over the maximum {count} times')
            print(f'Adjusting profiles...')
        final_profiles = create_profiles(reduce_dict)
        flag, count, new_reduce_dict = check_maximum_load(
            adjusted=True)  # check is maximum load is not higher than certain value (100kVA for phase II)
        reduce_dict = merge_dicts([reduce_dict, new_reduce_dict])
        if flag == 0:
            if print_updates:
                print('Sum of loads is under the maximum at all times')
        if flag == 1:
            if print_updates:
                print(f'Sum of loads is over the maximum {count} times')
                print(f'Adjusting profiles...')
            final_profiles = create_profiles(reduce_dict)
            flag, count, new_reduce_dict = check_maximum_load(
                adjusted=True)  # check is maximum load is not higher than certain value (100kVA for phase II)
            reduce_dict = merge_dicts([reduce_dict, new_reduce_dict])
            if flag == 0:
                if print_updates:
                    print('Sum of loads is under the maximum at all times')
            if flag == 1:
                if print_updates:
                    print(f'Sum of loads is over the maximum {count} times')

            return final_profiles

try:
    loads = pd.read_csv(os.path.join(os.getcwd(), 'Load.csv'), sep=';', index_col='id')
    profiles = pd.read_csv(os.path.join(os.getcwd(), 'LoadProfile.csv'), sep=';', index_col='time')
    pv_profiles = load_workbook(r"PV_Profiles.xlsx")
except FileNotFoundError:
    profiles_path = os.path.join('raw_data_generation', 'input', 'ERIGrid_Profiles_phase2')
    loads = pd.read_csv(os.path.join(profiles_path, 'Load.csv'), sep=';', index_col='id')
    profiles = pd.read_csv(os.path.join(profiles_path, 'LoadProfile.csv'), sep=';', index_col='time')
    pv_profiles = load_workbook(os.path.join(os.path.join(profiles_path, r"PV_Profiles.xlsx")))


if __name__ == '__main__':

    create_PV_profiles()
    create_profiles()

    finalize_profiles()






